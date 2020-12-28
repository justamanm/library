# pip install openpyxl
from openpyxl import load_workbook, Workbook


class ReadExcel:
    # 打开文件
    def open(self):
        self.wb = load_workbook('input.xlsx')
        # 获取sheet：
        self.sheet = self.wb.get_sheet_by_name('Sheet1')  # 通过表名获取

    # 属性和方法
    def attribut(self):
        # 获取行数和列数：
        row_num = self.sheet.max_row  # 获取行数
        col_num = self.sheet.max_column  # 获取列数
        columns = self.sheet.columns    # 每一列都是一个元组，各列组成大的元组
        rows = self.sheet.rows      # 每一行都是一个元组，各行组成大的元组

    # 获取指定单元格值，可以通过方法或索引
    def read_cell(self):
        # 通过方法
        row = 1
        col = 1
        data = self.sheet.cell(row=row, column=col).value  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value

        # 通过索引
        data = self.sheet["A1"]
        data = self.sheet["A1":"B1"]
        print(data)

    # 读取一行
    def read_row(self):
        self.open()

        # 通过方法，返回行的生成器
        rows = self.sheet.iter_rows(min_row=1, min_col=1)
        # 通过next获取一行
        row_01 = next(rows)
        print(row_01[:3])

        # 遍历获取每一行
        for i, row in enumerate(rows):
            row_cells = row[0]
            print(row_cells)

        # 通过索引
        row_01 = self.sheet[1]  # 第一行
        row_01_03 = self.sheet[1:3]     # 第1-3行
        print(row_01_03)

    # 读取一列
    def read_col(self):
        # 通过方法，返回列的生成器
        cols = self.sheet.iter_cols(min_col=1, min_row=1)
        for i, col in enumerate(cols):
            if i == 3:
                break
            col_cells = col[:3]
            print(col_cells)

        # 通过索引
        col_01 = self.sheet["A"]
        col_02 = self.sheet["B"]

    def delete(self):
        # 从第一列开始删，总共删2列，即删除1、2列
        self.sheet.delete_cols(1, 2)
        print("删除第一列")
        self.sheet.delete_rows(1)
        print("再删除第一行")
        self.save()

    def save(self):
        # 保存更改后的excel
        self.wb.save("test.xlsx")
        print("保存到data.xlsx")

    def run(self):
        self.open()
        # self.read_cell()
        # self.read_row()
        # self.read_col()
        # self.delete()
        self.save()


class WriteExcel:
    def create(self):
        self.wb = Workbook()
        # 创建sheet并指定位置
        self.sheet = self.wb.create_sheet("sheet01", 0)
        self.sheet_apply_empty = self.wb.create_sheet("sheet02", 1)

    # 在末尾增加
    def append(self):
        self.sheet.append(["第1行1列", "第1行2列", "第1行3列"])
        self.sheet.append(["第2行1列", "第2行2列", "第2行3列"])

        # 保存
        self.wb.save('data.xlsx')
        self.wb.close()

    def write(self):
        """
        修改某一列所有的单元格
        获取单元格位置，获取旧内容，写入新内容
        :return:
        """
        col_num = 2
        rows = self.sheet.max_row
        for row_num in range(3, rows + 1):
            print(f"处理第{col_num}列的第{row_num}行")
            # 获取旧内容
            old_value = self.sheet.cell(row=row_num, column=col_num).value
            # 做处理
            new_value = old_value.replace("a", "b")
            # 往单元格中写入处理后的新字符串内容，row指定行，column指定列
            self.sheet.cell(row=row_num, column=col_num).value = new_value  # if oldstr != "None" else ""
        self.save()

    def insert(self):
        # 在第二列前面插入空列，共插入两列
        self.sheet.insert_cols(2, 2)
        self.sheet.insert_rows(1)

    def delete(self):
        # 从第一列开始删，总共删2列，即删除1、2列
        self.sheet.delete_cols(1, 2)
        self.sheet.delete_rows(1)
        print("再删除第一行")
        self.save()

    def save(self):
        # 保存更改后的excel
        self.wb.save("data.xlsx")
        print("保存到data.xlsx")

    def run(self):
        pass
        # self.save()


if __name__ == '__main__':
    read = ReadExcel()
    read.run()

    write = WriteExcel()
    # write.run()